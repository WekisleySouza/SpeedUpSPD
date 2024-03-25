from openpyxl import Workbook, load_workbook

class Report:
    def __init__(self, times, name) -> None:
        self.__times = times
        self.__file_name = name
        self.__wb = Workbook()

    def __save_file(self) -> None:
        self.__wb.save(f'./reports/{self.__file_name}_times.xlsx')

    def __generate_report_base(self) -> Workbook:
        ws = self.__wb.active
        ws['A1'] = 'Tempos'
        ws['B1'] = 'Tempo médio'
        return ws

    def generate_times_report(self) -> None:
        base = self.__generate_report_base()
        
        for i, time in enumerate(self.__times):
            base[f'A{i+2}'] = f'{time:.2e}'

        base['B2'] = f'{sum(self.__times) / len(self.__times):.2e}'

        self.__save_file()

    def __generate_speedup_base() -> tuple:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Sequencial / 2 Threads'
        ws['A2'] = 'Sequencial / 4 Threads'
        ws['A3'] = 'Sequencial / 8 Threads'
        return wb, ws

    def generate_speedup_report() -> None:
        # Carregando as médias
        paralell_2_thread_average = load_workbook('./reports/teste_paralelo_2_threads_times.xlsx').active['B2'].value
        paralell_4_thread_average = load_workbook('./reports/teste_paralelo_4_threads_times.xlsx').active['B2'].value
        paralell_8_thread_average = load_workbook('./reports/teste_paralelo_8_threads_times.xlsx').active['B2'].value
        sequencial_average = load_workbook('./reports/teste_sequencial_times.xlsx').active['B2'].value
        # Escrevendo um novo arquivo com os speedups
        wb, ws = Report.__generate_speedup_base()
        ws['B1'] = float(sequencial_average) / float(paralell_2_thread_average)
        ws['B2'] = float(sequencial_average) / float(paralell_4_thread_average)
        ws['B3'] = float(sequencial_average) / float(paralell_8_thread_average)
        wb.save(f'./reports/speedup_tests.xlsx')